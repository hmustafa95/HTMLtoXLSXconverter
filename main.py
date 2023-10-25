import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup


# Function to open a file dialog and select an HTML file
def choose_html_file():
    html_file = filedialog.askopenfilename(filetypes=[("HTML files", "*.html")])
    html_file_entry.delete(0, tk.END)
    html_file_entry.insert(0, html_file)


# Function to skip appending question/answer to Excel file as it is unnecessary
def skip_questions(question):
    questions_to_skip = ['First name hidden', 'Surname hidden', 'Client email hidden', 'ID-Number', ]
    if question in questions_to_skip:
        return True


# Function to append Personal Information before the initial data
def personal_information(question):
    question_value = "Year"
    if question in question_value:
        row_personal_info = ["Personal", "Information"]
        return row_personal_info


# Function to append a blank row, Residency Information and a row for residency status before the residency data
def residency_information(question):
    question_value = 'Have you lived outside of the country in the last 3 years?'
    if question == question_value:
        blank_row = ['', '']
        resident_info = ["Residency", "Information"]
        residency_info = ["Resident", ""]
        return [blank_row, resident_info, residency_info]


# Function to append a blank row and Employment Income before the Employment data
def employment_information(question):
    question_value = 'How many employers did you have?'
    if question == question_value:
        blank_row = ['', '']
        employment_info = ['Employment', 'Income']
        return [blank_row, employment_info]


# Function to append a blank row and Self-Assessed Income before the SA data
def sa_information(question):
    question_value = 'Did you have any additional sources of income?'
    if question == question_value:
        blank_row = ['', '']
        sa_info = ['Self-Assessed', 'Income']
        return [blank_row, sa_info]


# Function to append a blank row and Personal Expenses before the Expenses data
def expenses_information(question):
    question_value = 'Did you pay rent during the tax year?'
    if question == question_value:
        blank_row = ['', '']
        tcs_info = ['Personal', 'Expenses']
        return [blank_row, tcs_info]


# Function to process the selected HTML file and save data to an Excel file
def process_html_to_excel():
    html_file = html_file_entry.get()
    excel_file = excel_file_entry.get()

    # Create a Font object to make the text bold
    bold_font = Font(bold=True)

    # Create a PatternFill object for a light blue background
    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Automatically append ".xlsx" extension if not provided by the user
    if not excel_file.lower().endswith('.xlsx'):
        excel_file += '.xlsx'

    try:
        # Try to open and read the HTML file
        with open(html_file, 'r', encoding='utf-8') as file:
            content = file.read()
    except FileNotFoundError:
        # Handle the case where the HTML file is not found
        result_label.config(text="Error: The HTML file not found.")
        return

    # Parse the HTML content using BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')

    try:
        # Try to open an existing Excel file and select the active worksheet
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active
    except FileNotFoundError:
        # Handle the case where the Excel file is not found, create a new file and select the active worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

    # Find the table in the HTML and extract data
    table = soup.find('table')
    if table:
        for row in table.find_all('tr'):
            question_cell = row.find('td')
            if question_cell:
                # Get the text from the question cell
                question_text = question_cell.get_text().strip()
                # Skip questions which are selected manually as they are repeated or unnecessary
                skip_question = skip_questions(question_text)
                if skip_question is True:
                    continue
                answer_cell = question_cell.find_next('td')
                # Get the value from the answer cell
                input_tag = answer_cell.find('input')

                residence_func = residency_information(question_text)
                if residence_func is not None:
                    for each_value in residence_func:
                        worksheet.append(each_value)

                        if each_value == ["Residency", "Information"]:
                            # Apply formatting to specific cells in the specific row
                            for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row,
                                                            min_col=1,
                                                            max_col=len(residence_func)):
                                for cell, value in zip(cell, each_value):
                                    cell.value = value
                                    cell.font = bold_font
                                    cell.fill = light_blue_fill

                paye_func = employment_information(question_text)
                if paye_func is not None:
                    for each_value in paye_func:
                        worksheet.append(each_value)

                        if each_value == ['Employment', 'Income']:
                            for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row,
                                                            min_col=1,
                                                            max_col=len(paye_func)):
                                for cell, value in zip(cell, each_value):
                                    cell.value = value
                                    cell.font = bold_font
                                    cell.fill = light_blue_fill

                sa_func = sa_information(question_text)
                if sa_func is not None:
                    for each_value in sa_func:
                        worksheet.append(each_value)

                        if each_value == ['Self-Assessed', 'Income']:
                            for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row,
                                                            min_col=1,
                                                            max_col=len(sa_func)):
                                for cell, value in zip(cell, each_value):
                                    cell.value = value
                                    cell.font = bold_font
                                    cell.fill = light_blue_fill

                tcs_func = expenses_information(question_text)
                if tcs_func is not None:
                    for each_value in tcs_func:
                        worksheet.append(each_value)

                        if each_value == ['Personal', 'Expenses']:
                            for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row,
                                                            min_col=1,
                                                            max_col=len(tcs_func)):
                                for cell, value in zip(cell, each_value):
                                    cell.value = value
                                    cell.font = bold_font
                                    cell.fill = light_blue_fill

                if input_tag:
                    input_value = input_tag.get('value')
                    if input_value != "No" and question_text not in unique_questions_texts:
                        personal_func = personal_information(question_text)
                        if personal_func is not None:
                            worksheet.append(personal_func)

                            for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=1,
                                                            max_col=len(personal_func)):
                                for cell, value in zip(cell, personal_func):
                                    cell.value = value
                                    cell.font = bold_font
                                    cell.fill = light_blue_fill

                        if input_value.isnumeric():
                            input_value = int(input_value)  # Convert numeric text to int
                        elif input_value.replace(".", "", 1).isdigit():
                            input_value = float(input_value)  # Convert numeric text to float

                        # Append question and answer to the Excel worksheet
                        worksheet.append([question_text, input_value])
                        
                        # Add the question_text to the set of unique questions
                        unique_questions_texts.add(question_text)

    # Save the Excel file
    workbook.save(excel_file)
    result_label.config(text=f"Data from HTML file appended to {excel_file} successfully.")


unique_questions_texts = set()

# Create the GUI
root = tk.Tk()
root.title("HTML to Excel Converter")

# Label and input field to select an HTML file
tk.Label(root, text="Select HTML File:").pack()
html_file_entry = tk.Entry(root)
html_file_entry.pack()
tk.Button(root, text="Browse", command=choose_html_file).pack()

# Label and input field to enter the Excel file name
tk.Label(root, text="Enter Excel File Name:").pack()
excel_file_entry = tk.Entry(root)
excel_file_entry.pack()

# Button to initiate the conversion process
tk.Button(root, text="Convert", command=process_html_to_excel).pack()

# Label to display the result of the conversion
result_label = tk.Label(root, text="")
result_label.pack()

# Start the Tkinter main loop
root.mainloop()
